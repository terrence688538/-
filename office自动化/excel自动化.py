import xlrd
xlrd=xlrd.open_workbook(r'D:\资源\python自动化\Python办公自动化视频\章节01：购后必读，学员福利\02-04 课件\CourseCode\Chapter1\S1-1-1\LessonCode\7月下旬入库表.xlsx')
table=xlrd.sheet_by_index(0)                        #根据位置选择表
#table_2=xlrd.sheet_by_name('七月下旬入库表')        #根据表名选择表
print(table.cell_value(1,2))
print(table.cell(1,2).value)
import xlwt                         #生成的列数不能超过256
new_workbook=xlwt.Workbook()
worksheet=new_workbook.add_sheet('new_test')
worksheet.write(0,0,'test')
new_workbook.save(r'D:\资源\python自动化\Python办公自动化视频\章节01：购后必读，学员福利\02-04 课件\CourseCode\Chapter1\S1-1-1\LessonCode\test.xls')
from xlutils.copy import copy
tem_excel=xlrd.open_workbook(r'D:\资源\python自动化\Python办公自动化视频\章节01：购后必读，学员福利\02-04 课件\CourseCode\Chapter1\S1-1-2\LessonCode\日统计.xls',formatting_innfo=True)
tem_sheet=tem_excel.sheet_by_index(0)
new_excel=copy(tem_excel)
new_sheet=new_excel.get_sheet(0)
style=xlwt.XFStyle()
font=xlwt.Font()
font.name='微软雅黑'
font.bold=True
font.height=360
style.font=font
#边框
borders=xlwt.Borders()
borders.top=xlwt.Borders.THIN
borders.right=xlwt.Borders.THIN
borders.left=xlwt.Borders.THIN
borders.bottom=xlwt.Borders.THIN
style.borders=borders
#对齐
aligment=xlwt.Alignment()
aligment.horz=xlwt.Alignment.HORZ_CENTER           #水平方向
aligment.vert=xlwt.Alignment.VERT_CENTER            #垂直方向
style.alignment=aligment

new_sheet.write(2,1,12,style)
new_sheet.write(3,1,12,style)
new_sheet.write(4,1,12,style)
new_sheet.write(5,1,12,style)
new_excel.save()

import pymysql
database=pymysql.connect('localhost','root','688538','girls')
cursor=database.cursor()           #初始化指针
sql="INSERT INTO beauty(`id`,`name`,`sex`,`borndate`,`phone`,`photo`,`boyfriend_id`) VALUES(21,'刘ii','女','1990-4-23','18888888',NULL,2);" #引号中是sql语句
cursor.execute(sql)    #执行
database.commit()      #对储存的数据修改后，需要commit
database.close()       #关掉

#关于查询
database=pymysql.connect('localhost','root','688538','myemployees')
cursor=database.cursor()
sql="SELECT * FROM employees WHERE last_name LIKE '%a%';"
cursor.execute(sql)
result=cursor.fetchall()

import xlsxwriter as xw            #支持超过256列     但不能带格式
workbook=xw.Workbook()
sheet0=workbook.add_worksheet('sheet0')

import openpyxl                    #性能不是很稳定
workbook=openpyxl.load_workbook(r'D:\资源\python自动化\Python办公自动化视频\章节01：购后必读，学员福利\02-04 课件\CourseCode\Chapter1\S1-1-1\LessonCode\7月下旬入库表.xlsx')
sheet0=workbook['Sheet1']
sheet0['B3']='5'
workbook.save()